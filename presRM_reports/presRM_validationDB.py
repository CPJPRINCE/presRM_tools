import pandas as pd
from pyPreservica import *
from .. import secret
from datetime import datetime 
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

#print(pyPreservica.__file__)

reportlocation = r"C:\Users\Chris.Prince\Unilever\UARM Teamsite - Records Management\Preservica Spreadsheet Uploads\TopLevelDB"

content = ContentAPI(username=secret.username,password=secret.password, \
                    tenant="UARM",server="unilever.preservica.com")

entity = EntityAPI(username=secret.username,password=secret.password, \
                    tenant="UARM",server="unilever.preservica.com")

starttime = datetime.now()
print(f"Started: {starttime}")
top_ref = "bf614a5d-10db-4ff8-addb-2c913b3149eb"
filters = {"xip.reference":"*","xip.identifier":"code","xip.parent_hierarchy":top_ref}
c = 0
folder = entity.folder(top_ref)
toplist = []
deptdf = pd.DataFrame()
for hit in entity.descendants(folder):
    toplist.append({"Top-Level Folders": hit.title,"Ref": hit.reference})
toplist = sorted(toplist,key=lambda d: d["Top-Level Folders"])
topdf = pd.DataFrame(toplist)
for hit in toplist:
    deptlist = []
    #print(hit)
    nfolder = entity.folder(hit.get('Ref'))
    for nhit in entity.descendants(nfolder):
        deptlist.append(nhit.title)
    dictdept = {f"{hit.get('Top-Level Folders')}":sorted(deptlist)}   
    ndf = pd.DataFrame(dictdept)
    deptdf = pd.concat([deptdf,ndf],axis=1)
file_output = os.path.join(reportlocation, "TopLevelDB.xlsx") 
with pd.ExcelWriter(file_output,'auto',mode='w') as writer:
    topdf.to_excel(writer,sheet_name="Top-VAL",index=False)
    deptdf.to_excel(writer,sheet_name="Dept-VAL",index=False)

wb = load_workbook(file_output)
ws = wb['Dept-VAL']
for count,column in enumerate(ws.columns,1):
    max_row = max(cell.row for cell in column if cell.value is not None)
    letter = get_column_letter(count)
    #print(f"{letter}1:{letter}{max_row}")
    table_name = str(column[0].value)
    #print(table_name)
    if "Legal Agreements and IT" in table_name: table_name = "LEGIT"
    elif "Market Clusters Organisation" in table_name: table_name = "MCO"
    else: table_name = table_name[0:3].upper()
    #print(table_name)
    table = Table(displayName=f"{table_name}", ref=f"{letter}1:{letter}{max_row}")
    style = TableStyleInfo(name="TableStyleMedium9",showFirstColumn=False,
                       showLastColumn=False,showRowStripes=True,showColumnStripes=False)
    table.tableStyleInfo = style
    ws.add_table(table)
topws = wb['Top-VAL']
ref=f"A1:{get_column_letter(topws.max_column) + str(topws.max_row)}"
print(ref)
table = Table(displayName="Top",ref=f"A1:{get_column_letter(topws.max_column) + str(topws.max_row)}")
table.tableStyleInfo = style
topws.add_table(table)
wb.save(file_output)

print(f"Complete! Ran for: {datetime.now() - starttime}")